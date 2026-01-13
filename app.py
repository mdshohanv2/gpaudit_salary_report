import streamlit as st
import pandas as pd
import io

def main():
    st.set_page_config(layout="wide") # Set page layout to wide for better use of space
    st.title("Auditor Performance and Salary Analysis")

    # Custom CSS for table alignment and potentially font size (re-adding some for general readability)
    st.markdown("""
        <style>
            /* Left-align the table */
            table {
                margin-left: 0 !important;
                margin-right: auto !important;
            }
            /* General font size for the table content */
            .dataframe {
                font-size: 1.1em; 
            }
            .dataframe th, .dataframe td {
                padding: 8px 10px; /* Adjust padding for better spacing */
            }
        </style>
        """, unsafe_allow_html=True)

    st.sidebar.header("Settings & Uploads")
    unit_price = st.sidebar.number_input("Unit Price (BDT)", min_value=0, value=3, step=1)
    
    st.sidebar.markdown("---")
    audit_file = st.sidebar.file_uploader("Upload Audit Data (CSV or XLSX)", type=["csv", "xlsx"])
    mfs_file = st.sidebar.file_uploader("Upload MFS Data (CSV)", type=["csv"])

    if audit_file is not None and mfs_file is not None:
        try:
            # --- Process Audit Data ---
            if audit_file.name.endswith('.csv'):
                df_audit = pd.read_csv(audit_file)
            else: # .xlsx
                df_audit = pd.read_excel(audit_file)

            # Ensure 're_audited' and 'mismatch_found_in_reaudit' are boolean
            df_audit['re_audited'] = df_audit['re_audited'].astype(bool)
            df_audit['mismatch_found_in_reaudit'] = df_audit['mismatch_found_in_reaudit'].astype(bool)

            # --- Extract Date Info for Header ---
            # Search for a column with 'date' in its name
            date_col = next((col for col in df_audit.columns if 'date' in col.lower()), None)
            header_title = "GP GLM Auditor's Salary"
            header_date_range = "Visit Date: [Date Not Found]"
            
            if date_col:
                # Attempt to parse dates
                try:
                    df_audit[date_col] = pd.to_datetime(df_audit[date_col])
                    valid_dates = df_audit[date_col].dropna()
                    if not valid_dates.empty:
                        start_date = valid_dates.min()
                        end_date = valid_dates.max()
                        
                        # Format: dd-Month-yyyy
                        start_str = start_date.strftime('%d-%B-%Y')
                        end_str = end_date.strftime('%d-%B-%Y')
                        
                        header_date_range = f"Visit Date: {start_str} to {end_str}"
                        # Salary month based on end date: e.g., December'2025
                        header_title = f"GP GLM Auditor's Salary- {end_date.strftime('%B')}'{end_date.year}"
                except Exception:
                    pass

            # Group by auditor name (assigned_to) for audit performance
            auditor_performance = df_audit.groupby('assigned_to').apply(lambda group: pd.Series({
                'audit_visited': group['visit_id'].nunique(),
                're_audit_visited': group['re_audited'].sum(),
                'mismatch_found_no_audit': group[group['re_audited'] == True]['mismatch_found_in_reaudit'].eq(False).sum(),
                'mismatch_found_yes_audit': group[group['re_audited'] == True]['mismatch_found_in_reaudit'].eq(True).sum()
            })).reset_index()

            # Calculate % for logic (keep as float)
            auditor_performance['mismatch_rate'] = (
                auditor_performance['mismatch_found_yes_audit'] / auditor_performance['re_audit_visited']
            ).fillna(0)

            # Performance Calculations
            auditor_performance['Unit Price'] = unit_price
            auditor_performance['Max Payable'] = auditor_performance['audit_visited'] * unit_price
            auditor_performance['Fixed (75%)'] = auditor_performance['Max Payable'] * 0.75
            auditor_performance['Variable (25%)'] = (auditor_performance['Max Payable'] * 0.25) * (1 - auditor_performance['mismatch_rate'])
            auditor_performance['Actual Payable'] = auditor_performance['Fixed (75%)'] + auditor_performance['Variable (25%)']
            
            # Final Percentage for display
            auditor_performance['% Mismatch in Re-Audit'] = auditor_performance['mismatch_rate'] * 100

            # Rename columns for display
            auditor_performance.rename(columns={
                'assigned_to': 'Auditor Name',
                'audit_visited': 'Audited Visit',
                're_audit_visited': 'Re-Audited Visit',
                'mismatch_found_no_audit': 'Mismatch No',
                'mismatch_found_yes_audit': 'Mismatch Yes'
            }, inplace=True)

            # --- Process MFS Data ---
            df_mfs = pd.read_csv(mfs_file, header=2)
            mfs_data = df_mfs[['Auditor Name', 'Full Name', 'MFS Number', 'MFS Provider']].copy()
            
            # Ensure MFS Number starts with 0
            mfs_data['MFS Number'] = mfs_data['MFS Number'].apply(lambda x: f"0{int(x)}" if pd.notnull(x) and not str(x).startswith('0') else str(x))

            # --- Merge Data ---
            combined_df = pd.merge(
                auditor_performance,
                mfs_data,
                on='Auditor Name',
                how='left'
            )

            # Insert 'Sl' column at the beginning
            combined_df.insert(0, 'Sl', range(1, len(combined_df) + 1))

            # Add Grand Total row
            numeric_cols = ['Audited Visit', 'Re-Audited Visit', 'Mismatch No', 'Mismatch Yes', 'Max Payable', 'Fixed (75%)', 'Variable (25%)', 'Actual Payable']
            totals = combined_df[numeric_cols].sum()
            
            total_row = pd.DataFrame([{
                'Auditor Name': 'GRAND TOTAL',
                **totals.to_dict(),
                '% Mismatch in Re-Audit': (totals['Mismatch Yes'] / totals['Re-Audited Visit'] * 100) if totals['Re-Audited Visit'] > 0 else 0,
                'Unit Price': ''
            }])
            df_total_row = pd.concat([combined_df, total_row], ignore_index=True)

            # --- Prepare Data for Excel Formulas ---
            # We keep a copy of the dataframe before string formatting
            excel_df = df_total_row.copy()
            
            # Format columns for Frontend display (using original name combined_df for display logic)
            combined_df = df_total_row.copy()

            # Format columns for Frontend
            if '% Mismatch in Re-Audit' in combined_df.columns:
                combined_df['% Mismatch in Re-Audit'] = combined_df['% Mismatch in Re-Audit'].apply(
                    lambda x: f"{round(float(x))}%" if pd.notnull(x) and x != '' else ""
                )

            # Round numeric payment columns to whole numbers for display
            payment_cols = ['Max Payable', 'Fixed (75%)', 'Variable (25%)', 'Actual Payable']
            for col in payment_cols:
                combined_df[col] = combined_df[col].apply(lambda x: round(float(x)) if pd.notnull(x) and x != '' else x)

            cols_to_int = ['Sl', 'Audited Visit', 'Re-Audited Visit', 'Mismatch No', 'Mismatch Yes'] + payment_cols
            for col in cols_to_int:
                if col in combined_df.columns:
                    combined_df[col] = combined_df[col].fillna(0).astype(str).replace(r'\.0$', '', regex=True)
            
            combined_df.replace('0', '', inplace=True)
            combined_df.replace('nan', '', inplace=True)
            combined_df.loc[combined_df['Auditor Name'] == 'GRAND TOTAL', 'Sl'] = ''
            combined_df.fillna('', inplace=True)

            # Reorder columns to match image
            final_cols = [
                'Sl', 'Auditor Name', 'Audited Visit', 'Re-Audited Visit', 'Mismatch No', 'Mismatch Yes', 
                '% Mismatch in Re-Audit', 'Unit Price', 'Max Payable', 'Fixed (75%)', 'Variable (25%)', 
                'Actual Payable', 'Full Name', 'MFS Number', 'MFS Provider'
            ]
            combined_df = combined_df[final_cols]

            # Dynamic Header Display
            st.markdown(f"""
                <div style="text-align: center; border: 1px solid #e6e9ef; padding: 20px; background-color: #f8f9fb; border-radius: 10px; margin-bottom: 20px;">
                    <h1 style="margin: 0; color: #1f365d; font-size: 2em;">{header_title}</h1>
                    <p style="margin: 10px 0 0 0; font-size: 1.25em; color: #4a5568; font-weight: 500;">
                        [{header_date_range}]
                    </p>
                </div>
                """, unsafe_allow_html=True)
            # Using st.dataframe for interactive features like sorting, resizing, and cell selection
            st.dataframe(
                combined_df, 
                use_container_width=True, 
                hide_index=True,
                height=600
            )

            # Add download buttons in columns
            col1, col2 = st.columns(2)
            
            with col1:
                csv = combined_df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="Download as CSV",
                    data=csv,
                    file_name="combined_auditor_performance.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            
            with col2:
                # Excel export logic with high-fidelity styling
                excel_buffer = io.BytesIO()
                # Exclude 'Sl' for Excel to match the image
                export_cols = [
                    'Auditor Name', 'Audited Visit', 'Re-Audited Visit', 
                    'Mismatch No', 'Mismatch Yes', '% Mismatch in Re-Audit', 
                    'Unit Price', 'Max Payable', 'Fixed (75%)', 'Variable (25%)', 
                    'Actual Payable', 'Full Name', 'MFS Number', 'MFS Provider'
                ]
                
                with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
                    # Write static data starting from row 5 (Row 1-2: Title, Row 3-4: Headers)
                    excel_df[export_cols].to_excel(writer, index=False, sheet_name='Salary Report', startrow=4, header=False)
                    
                    workbook = writer.book
                    worksheet = writer.sheets['Salary Report']
                    
                    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
                    
                    # --- Color and Style Definitions ---
                    blue_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                    yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
                    white_font = Font(color='FFFFFF', bold=True, size=12)
                    black_bold_font = Font(bold=True)
                    center_aligned = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    thin_border = Border(
                        left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )

                    # --- 1. Main Headers (Rows 1 & 2) ---
                    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(export_cols))
                    worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(export_cols))
                    
                    row1 = worksheet.cell(row=1, column=1, value=header_title)
                    row2 = worksheet.cell(row=2, column=1, value=f"[{header_date_range}]")
                    
                    for r in [1, 2]:
                        for c in range(1, len(export_cols) + 1):
                            cell = worksheet.cell(row=r, column=c)
                            cell.fill = blue_fill
                            cell.font = white_font
                            cell.alignment = center_aligned

                    # --- 2. Table Headers (Rows 3 & 4) ---
                    # Columns: A(Auditor), B(Audited), C(ReAudited), D-F(Mismatch), G(Unit), H(Max), I(Fixed), J(Variable), K(Actual), L(Full), M(MFS), N(Provider)
                    
                    headers = [
                        (3, 1, 4, 1, "Auditor Name"),
                        (3, 2, 4, 2, "Audited Visit"),
                        (3, 3, 4, 3, "Re-Audited\nVisit"),
                        (3, 4, 3, 6, "Mismatch found"), # Parent for No, Yes, %
                        (4, 4, 4, 4, "No"),
                        (4, 5, 4, 5, "Yes"),
                        (4, 6, 4, 6, "%"),
                        (3, 7, 4, 7, "Unit Price"),
                        (3, 8, 4, 8, "Max Payable"),
                        (3, 9, 3, 9, "Fixed"),
                        (4, 9, 4, 9, "75%"),
                        (3, 10, 3, 10, "Variable"),
                        (4, 10, 4, 10, "25%"),
                        (3, 11, 4, 11, "Actual\nPayable"),
                        (3, 12, 4, 12, "Full Name"),
                        (3, 13, 4, 13, "MFS Number"),
                        (3, 14, 4, 14, "MFS Provider"),
                    ]

                    for s_row, s_col, e_row, e_col, val in headers:
                        if s_row != e_row or s_col != e_col:
                            worksheet.merge_cells(start_row=s_row, start_column=s_col, end_row=e_row, end_column=e_col)
                        cell = worksheet.cell(row=s_row, column=s_col, value=val)
                        cell.font = black_bold_font
                        cell.alignment = center_aligned
                        # Apply borders to all cells in the header area
                        for r in range(s_row, e_row + 1):
                            for c in range(s_col, e_col + 1):
                                worksheet.cell(row=r, column=c).border = thin_border

                    # --- 3. Data Rows & Formulas ---
                    num_auditors = len(excel_df) - 1
                    data_start_row = 5
                    
                    for i in range(num_auditors + 1): # +1 to include Grand Total row
                        row_idx = data_start_row + i
                        is_total = (i == num_auditors)
                        
                        # Apply borders and alignment to row
                        for col_idx in range(1, len(export_cols) + 1):
                            cell = worksheet.cell(row=row_idx, column=col_idx)
                            cell.border = thin_border
                            cell.alignment = center_aligned if col_idx > 1 else Alignment(horizontal='left')
                            
                            # Actual Payable Column (K / 11) gets yellow background
                            if col_idx == 11:
                                cell.fill = yellow_fill
                                cell.font = black_bold_font

                        if not is_total:
                            # Formulas for auditor rows
                            # % Mismatch (E/D): worksheet.cell(row=row_idx, column=6).value
                            # Note: Column indices: B=2(Audited), C=3(ReAudited), D=4(No), E=5(Yes), F=6(%), G=7(Unit), H=8(Max), I=9(Fixed), J=10(Variable), K=11(Actual)
                            worksheet.cell(row=row_idx, column=6).value = f"=IF(D{row_idx}=0, 0, E{row_idx}/D{row_idx})"
                            worksheet.cell(row=row_idx, column=6).number_format = '0.00%'
                            
                            # Max Payable (B*G) - Rounded to 0
                            worksheet.cell(row=row_idx, column=8).value = f"=ROUND(B{row_idx}*G{row_idx}, 0)"
                            # Fixed (H*0.75) - Rounded to 0
                            worksheet.cell(row=row_idx, column=9).value = f"=ROUND(H{row_idx}*0.75, 0)"
                            # Variable ((H*0.25)*(1-F)) - Rounded to 0
                            worksheet.cell(row=row_idx, column=10).value = f"=ROUND((H{row_idx}*0.25)*(1-F{row_idx}), 0)"
                            # Actual (I+J)
                            worksheet.cell(row=row_idx, column=11).value = f"=I{row_idx}+J{row_idx}"
                            
                            # Apply number format to payment columns
                            for pay_col in [8, 9, 10, 11]:
                                worksheet.cell(row=row_idx, column=pay_col).number_format = '#,##0'
                        else:
                            # Grand Total formula row
                            for col_letter in ['B', 'C', 'D', 'E', 'H', 'I', 'J', 'K']:
                                worksheet[f"{col_letter}{row_idx}"] = f"=SUM({col_letter}{data_start_row}:{col_letter}{row_idx - 1})"
                                if col_letter in ['H', 'I', 'J', 'K']:
                                    worksheet[f"{col_letter}{row_idx}"].number_format = '#,##0'
                            
                            # Total %: =IF(D=0, 0, E/D)
                            worksheet[f"F{row_idx}"] = f"=IF(D{row_idx}=0, 0, E{row_idx}/D{row_idx})"
                            worksheet[f"F{row_idx}"].number_format = '0.00%'

                    # Set column widths for better visibility
                    column_widths = [25, 12, 12, 8, 8, 10, 10, 12, 10, 10, 12, 25, 15, 15]
                    for i, width in enumerate(column_widths):
                        worksheet.column_dimensions[chr(65 + i)].width = width

                st.download_button(
                    label="Download as Excel with Formulas",
                    data=excel_buffer.getvalue(),
                    file_name=f"Salary_Report_{header_title.split('- ')[-1]}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

        except Exception as e:
            st.error(f"An error occurred during file processing: {e}")
    else:
        st.info("Please upload both audit data and MFS data files to see the analysis.")

if __name__ == "__main__":
    main()